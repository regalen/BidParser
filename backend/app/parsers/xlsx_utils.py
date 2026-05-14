from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.parsers.base import ParseError
from app.parsers.cleaning import clean_text, parse_decimal


def active_sheet(path: Path) -> Worksheet:
    workbook = load_workbook(path, data_only=True, read_only=False)
    return workbook.active


def find_cell(ws: Worksheet, value: str, *, min_row: int = 1) -> tuple[int, int]:
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            if clean_text(cell.value) == value:
                return cell.row, cell.column
    raise ParseError(f"Could not find anchor cell {value!r}", stage="detect", hint=f"Could not find {value!r}.")


def find_cell_starting(ws: Worksheet, prefix: str, *, min_row: int = 1) -> tuple[int, int, str]:
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            text = clean_text(cell.value)
            if text.startswith(prefix):
                return cell.row, cell.column, text
    raise ParseError(f"Could not find cell starting with {prefix!r}", stage="parse", hint=f"Could not find {prefix!r}.")


def header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    labels: dict[str, int] = {}
    for cell in ws[header_row]:
        text = clean_text(cell.value)
        if text:
            labels[text] = cell.column
    return labels


def require_labels(labels: dict[str, int], required: list[str]) -> None:
    missing = [label for label in required if label not in labels]
    if missing:
        raise ParseError(
            f"Missing required header labels: {', '.join(missing)}",
            stage="parse",
            hint=f"Missing required columns: {', '.join(missing)}.",
        )


def row_is_empty(ws: Worksheet, row_number: int) -> bool:
    return all(clean_text(cell.value) == "" for cell in ws[row_number])


def cell_text(ws: Worksheet, row: int, column: int) -> str:
    return clean_text(ws.cell(row=row, column=column).value)


def cell_value(ws: Worksheet, row: int, column: int) -> Any:
    return ws.cell(row=row, column=column).value


def parse_total_text(text: str) -> Decimal:
    return parse_decimal(text.replace("TOTAL", "", 1))
