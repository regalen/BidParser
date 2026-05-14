from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from app.parsers.base import LineItem


CRM_TEMPLATE_BY_VENDOR = {
    "Nutanix": "Foreign Uplift",
}

HEADERS = [
    "Item",
    "Vendor Name",
    "IMTH SKU\n(Optional)",
    "Vendor Part Number",
    "Description",
    "Qty.",
    "Unit Price",
    "MSRP",
    "Cost",
    "Discount",
    "Margin",
    "Product Part Number \n(for Warranty/Renewal)",
    "Serial Number",
    "Warranty / Duration (months)",
    "Vendor Ref",
    "Start Date",
    "End Date",
    "Comments",
    "Foreign Currency",
    "Foreign Cost",
    "Foreign MSRP",
    "Foreign Exchange Rate",
    "Min Order Qty",
    "IM%",
    "Diff%",
    "On Cost %",
    "Retail Bump %",
]

END_LOOP_WARNING = "DO NOT DELETE THIS LINE. Indicate * on column B to mark the end loop. Add / remove lines above as necessary."


def output_filename(source_filename: str) -> str:
    return f"{Path(source_filename).stem}_parsed.xlsx"


def write_foreign_uplift(
    line_items: Iterable[LineItem],
    output_path: Path,
    *,
    margin: Decimal = Decimal("5.00"),
    fx_rate: Decimal = Decimal("1.000"),
    vendor_name: str = "NUTANIX",
    currency: str = "USD",
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Foreign Uplift"

    ws.cell(row=1, column=12, value="(Optional for Software and/or Services)")
    for index, header in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=index, value=header)

    row_number = 3
    for item_index, item in enumerate(line_items, start=1):
        ws.cell(row=row_number, column=1, value=item_index)
        ws.cell(row=row_number, column=2, value=vendor_name)
        ws.cell(row=row_number, column=4, value=item.vpn)
        ws.cell(row=row_number, column=5, value=item.description)
        ws.cell(row=row_number, column=6, value=item.qty)
        ws.cell(row=row_number, column=11, value=_excel_number(margin))
        if item.term is not None and item.term >= 1:
            ws.cell(row=row_number, column=14, value=item.term)
        if item.start_date is not None:
            cell = ws.cell(row=row_number, column=16, value=_excel_date(item.start_date))
            cell.number_format = "DD/MM/YYYY"
        if item.end_date is not None:
            cell = ws.cell(row=row_number, column=17, value=_excel_date(item.end_date))
            cell.number_format = "DD/MM/YYYY"
        ws.cell(row=row_number, column=18, value=item.serial_number)
        ws.cell(row=row_number, column=19, value=currency)
        ws.cell(row=row_number, column=20, value=_excel_number(item.cost))
        ws.cell(row=row_number, column=21, value=_excel_number(item.msrp if item.msrp is not None else Decimal("0")))
        ws.cell(row=row_number, column=22, value=_excel_number(fx_rate))
        row_number += 1

    ws.cell(row=row_number, column=2, value="*")
    ws.cell(row=row_number, column=4, value=END_LOOP_WARNING)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _excel_number(value: Decimal | int | float) -> int | float:
    decimal = value if isinstance(value, Decimal) else Decimal(str(value))
    if decimal == decimal.to_integral_value():
        return int(decimal)
    return float(decimal)


def _excel_date(value: date) -> datetime:
    return datetime(value.year, value.month, value.day)
