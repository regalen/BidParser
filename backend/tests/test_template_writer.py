from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.output.template_writer import write_foreign_uplift
from app.parsers.registry import get_parser


ROOT = Path(__file__).resolve().parents[2]
INPUTS = ROOT / "samples" / "inputs"
OUTPUTS = ROOT / "samples" / "outputs"


CASES = [
    ("nutanix_software_only_pdf", "XQ-4076249.pdf", "XQ-4076249_parsed.xlsx"),
    ("nutanix_software_only_xlsx", "XQ-4076249.xlsx", "XQ-4076249_parsed.xlsx"),
    ("nutanix_hardware_only_pdf", "XQ-4108785.pdf", "XQ-4108785_parsed.xlsx"),
    ("nutanix_hardware_only_xlsx", "XQ-4108785.xlsx", "XQ-4108785_parsed.xlsx"),
    ("nutanix_renewal_pdf", "XQ-4128926.pdf", "XQ-4128926_parsed.xlsx"),
]


@pytest.mark.parametrize(("slug", "input_name", "expected_name"), CASES)
def test_template_writer_matches_golden_workbook_cells(tmp_path: Path, slug: str, input_name: str, expected_name: str) -> None:
    result = get_parser(slug).parse(INPUTS / input_name)
    actual_path = tmp_path / expected_name

    write_foreign_uplift(result.line_items, actual_path, fx_rate=Decimal("1.000"), margin=Decimal("5.00"))

    assert_workbooks_equal(actual_path, OUTPUTS / expected_name)


def assert_workbooks_equal(actual_path: Path, expected_path: Path) -> None:
    actual = load_workbook(actual_path)
    expected = load_workbook(expected_path)

    assert actual.sheetnames == expected.sheetnames
    for sheet_name in expected.sheetnames:
        actual_ws = actual[sheet_name]
        expected_ws = expected[sheet_name]
        assert actual_ws.max_row == expected_ws.max_row
        assert actual_ws.max_column == expected_ws.max_column
        for row in range(1, expected_ws.max_row + 1):
            for column in range(1, expected_ws.max_column + 1):
                actual_cell = actual_ws.cell(row=row, column=column)
                expected_cell = expected_ws.cell(row=row, column=column)
                assert actual_cell.value == expected_cell.value, f"{sheet_name}!{actual_cell.coordinate}"
                assert actual_cell.number_format == expected_cell.number_format, f"{sheet_name}!{actual_cell.coordinate} number format"
